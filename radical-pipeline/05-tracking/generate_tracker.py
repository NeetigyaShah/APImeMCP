#!/usr/bin/env python3
"""Read tracker-data.json + merge every status/*.json -> APImeMCP-Radical-Tracker.xlsx.

Three sheets: Feature Catalog, Progress, Schedule/Deadlines.
Re-run any time (after each gate/wave update). Skips a malformed status file
gracefully (logs a warning, treats that feature as all-Todo) rather than crashing.

    pip install openpyxl
    python generate_tracker.py
"""
import json
import os
import glob
import datetime
import sys

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import DataBarRule
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

HERE = os.path.dirname(os.path.abspath(__file__))
DATA_FILE = os.path.join(HERE, "tracker-data.json")
STATUS_DIR = os.path.join(HERE, "status")
OUT_FILE = os.path.join(HERE, "APImeMCP-Radical-Tracker.xlsx")

S_KEYS = [f"S{i}" for i in range(12)]

# ---------- palette ----------
HEADER_FILL = PatternFill("solid", fgColor="1F2937")
HEADER_FONT = Font(color="FFFFFF", bold=True)
BORDER_THIN = Border(*(Side(style="thin", color="D1D5DB"),) * 4)

PROGRAM_FILL = {1: PatternFill("solid", fgColor="E0E7FF"), 2: PatternFill("solid", fgColor="DCFCE7")}

PILLAR_COLORS = [
    "FDE68A", "BFDBFE", "FBCFE8", "C7D2FE", "FECACA", "BBF7D0", "FED7AA",
    "A5F3FC", "DDD6FE",
]

RISK_FILL = {
    "Low": PatternFill("solid", fgColor="BBF7D0"),
    "Medium": PatternFill("solid", fgColor="FDE68A"),
    "High": PatternFill("solid", fgColor="FCA5A5"),
}

GATE_YES_FILL = PatternFill("solid", fgColor="93C5FD")
GATE_NO_FILL = PatternFill("solid", fgColor="F3F4F6")
CRITICAL_FILL = PatternFill("solid", fgColor="FCA5A5")
NONCRITICAL_FILL = PatternFill("solid", fgColor="F3F4F6")

SUBTASK_FILL = {
    "N/A": PatternFill("solid", fgColor="D1D5DB"),
    "Todo": PatternFill("solid", fgColor="FFFFFF"),
    "In-Prog": PatternFill("solid", fgColor="93C5FD"),
    "In-Review": PatternFill("solid", fgColor="FCD34D"),
    "Blocked": PatternFill("solid", fgColor="F87171"),
    "Done": PatternFill("solid", fgColor="86EFAC"),
}
SUBTASK_DONE_LIKE = {"Done", "N/A"}

STATUS_VS_PLAN_FILL = {
    "Ahead": PatternFill("solid", fgColor="86EFAC"),
    "On-track": PatternFill("solid", fgColor="BBF7D0"),
    "At-risk": PatternFill("solid", fgColor="FDE68A"),
    "Late": PatternFill("solid", fgColor="FCA5A5"),
}

WAVE_BAND_A = PatternFill("solid", fgColor="F9FAFB")
WAVE_BAND_B = PatternFill("solid", fgColor="EEF2FF")


def load_tracker_data():
    with open(DATA_FILE, encoding="utf-8") as f:
        return json.load(f)


def load_all_status():
    """id -> status dict. Malformed/missing files are skipped with a warning
    and synthesized as all-Todo so the sheet still renders."""
    statuses = {}
    for path in sorted(glob.glob(os.path.join(STATUS_DIR, "*.json"))):
        fid = os.path.splitext(os.path.basename(path))[0]
        try:
            with open(path, encoding="utf-8") as f:
                statuses[fid] = json.load(f)
        except (json.JSONDecodeError, OSError) as e:
            print(f"WARNING: skipping malformed status file {path}: {e}", file=sys.stderr)
    return statuses


def fallback_status(feature):
    subtasks = {k: ("Todo" if feature["subtasksApplicable"].get(k) else "N/A") for k in S_KEYS}
    return {
        "id": feature["id"], "subtasks": subtasks, "overall": "Not-started",
        "currentGate": None, "blockedBy": None, "owner": None, "reviewer": None, "updatedAt": "",
    }


def pillar_fill_map(features):
    pillars = sorted({f["pillar"] for f in features})
    return {p: PatternFill("solid", fgColor=PILLAR_COLORS[i % len(PILLAR_COLORS)]) for i, p in enumerate(pillars)}


def set_col_widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def style_header(ws, ncols, row=1):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER_THIN
    ws.freeze_panes = f"A{row + 1}"


def pct_complete(feature, status):
    applicable = [
        k for k in S_KEYS
        if feature["subtasksApplicable"].get(k)
        and status.get("subtasks", {}).get(k, "N/A") != "N/A"
    ]
    if not applicable:
        return 0.0
    done = sum(1 for k in applicable if status.get("subtasks", {}).get(k) in ("Done",))
    return done / len(applicable)


def build_catalog_sheet(wb, features, pillar_fills):
    ws = wb.active
    ws.title = "Feature Catalog"
    headers = [
        "ID", "Name", "Program", "Surface", "Pillar", "Description", "Value / Why",
        "Tool / Route / Screen added", "Primary modules", "New module?", "Skills",
        "Deps", "Wave", "Critical?", "Gate: Arch", "Gate: Security", "Gate: Live",
        "Owner", "Risk",
    ]
    ws.append(headers)
    style_header(ws, len(headers))

    for f in features:
        row = [
            f["id"], f["name"], f"Program {f['program']}", f["surface"], f["pillar"],
            f["description"], f["valueWhy"], f["toolOrSurface"], ", ".join(f["modules"]),
            "Yes" if f["newModule"] else "No", ", ".join(f["skills"]), ", ".join(f["deps"]) or "—",
            str(f["wave"]), "Yes" if f["criticalPath"] else "No",
            "Yes" if f["gates"]["arch"] else "No", "Yes" if f["gates"]["security"] else "No",
            "Yes" if f["gates"]["live"] else "No", f["owner"] or "—", f["risk"],
        ]
        ws.append(row)

    for r in range(2, ws.max_row + 1):
        f = features[r - 2]
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = BORDER_THIN
            cell.alignment = Alignment(vertical="top", wrap_text=(c in (6, 7, 8, 9, 11)))
        ws.cell(row=r, column=3).fill = PROGRAM_FILL[f["program"]]
        ws.cell(row=r, column=5).fill = pillar_fills[f["pillar"]]
        ws.cell(row=r, column=14).fill = CRITICAL_FILL if f["criticalPath"] else NONCRITICAL_FILL
        ws.cell(row=r, column=15).fill = GATE_YES_FILL if f["gates"]["arch"] else GATE_NO_FILL
        ws.cell(row=r, column=16).fill = GATE_YES_FILL if f["gates"]["security"] else GATE_NO_FILL
        ws.cell(row=r, column=17).fill = GATE_YES_FILL if f["gates"]["live"] else GATE_NO_FILL
        ws.cell(row=r, column=19).fill = RISK_FILL.get(f["risk"], GATE_NO_FILL)

    set_col_widths(ws, [7, 30, 10, 9, 13, 46, 46, 34, 30, 10, 34, 16, 6, 9, 9, 11, 9, 12, 9])
    ws.auto_filter.ref = ws.dimensions
    return ws


def build_progress_sheet(wb, features, statuses):
    ws = wb.create_sheet("Progress")
    headers = (
        ["ID", "Name", "Program"] + S_KEYS
        + ["% Complete", "Current gate", "Blocked by", "Owner", "Reviewer", "Overall status", "Last updated"]
    )
    ws.append(headers)
    style_header(ws, len(headers))

    pct_col = 3 + len(S_KEYS) + 1  # 1-indexed column of "% Complete"

    for f in features:
        st = statuses.get(f["id"]) or fallback_status(f)
        pct = pct_complete(f, st)
        row = (
            [f["id"], f["name"], f"Program {f['program']}"]
            + [st.get("subtasks", {}).get(k, "N/A") for k in S_KEYS]
            + [pct, st.get("currentGate") or "—", st.get("blockedBy") or "—",
               st.get("owner") or "—", st.get("reviewer") or "—",
               st.get("overall") or "Not-started", st.get("updatedAt") or "—"]
        )
        ws.append(row)

    for r in range(2, ws.max_row + 1):
        f = features[r - 2]
        st = statuses.get(f["id"]) or fallback_status(f)
        blocked = (st.get("overall") == "Blocked") or bool(st.get("blockedBy"))
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = BORDER_THIN
            if blocked and c <= 2:
                cell.fill = PatternFill("solid", fgColor="FEE2E2")
        for i, k in enumerate(S_KEYS):
            cell = ws.cell(row=r, column=4 + i)
            val = st.get("subtasks", {}).get(k, "N/A")
            cell.fill = SUBTASK_FILL.get(val, SUBTASK_FILL["N/A"])
            cell.alignment = Alignment(horizontal="center")
        ws.cell(row=r, column=pct_col).number_format = "0%"

    ws.conditional_formatting.add(
        f"{get_column_letter(pct_col)}2:{get_column_letter(pct_col)}{ws.max_row}",
        DataBarRule(start_type="num", start_value=0, end_type="num", end_value=1, color="22C55E"),
    )

    set_col_widths(ws, [7, 30, 10] + [4] * len(S_KEYS) + [11, 12, 16, 12, 12, 15, 20])
    ws.freeze_panes = "D2"  # freeze header row + ID/Name/Program columns
    ws.auto_filter.ref = ws.dimensions
    return ws


def add_days(date_str, days):
    if date_str is None:
        return None
    try:
        d = datetime.date.fromisoformat(date_str)
    except ValueError:
        return None
    return d + datetime.timedelta(days=days)


def wave_order(wave_ids):
    """Stable order for wave bands: engine 0..5 numerically, platform P0..P4 lexically after."""
    def key(w):
        s = str(w)
        if s.startswith("P"):
            return (1, int(s[1:]))
        return (0, int(s))
    return sorted(set(wave_ids), key=key)


def build_schedule_sheet(wb, features, statuses, tracker_data):
    ws = wb.create_sheet("Schedule-Deadlines")  # ponytail: "/" is illegal in xlsx sheet names
    headers = [
        "ID", "Name", "Program", "Wave", "Planned start", "Spec-done", "Build-complete",
        "Review-passed", "Live-verify", "Merged", "Promote", "Duration est (days)",
        "Status vs plan", "Owner",
    ]
    ws.append(headers)
    style_header(ws, len(headers))

    start_date_raw = tracker_data.get("START_DATE")
    start_date_set = False
    try:
        base_date = datetime.date.fromisoformat(start_date_raw)
        start_date_set = True
    except (TypeError, ValueError):
        base_date = None  # "TBD-set-at-execution" placeholder -> dates left blank

    durations = tracker_data.get("waveDurationEstimatesDays", {})
    waves_in_order = wave_order(f["wave"] for f in features)
    wave_start_offset = {}
    offset = 0
    for w in waves_in_order:
        wave_start_offset[w] = offset
        offset += durations.get(str(w), 7)

    milestone_fracs = {"spec": 0.15, "build": 0.55, "review": 0.75, "live": 0.9, "merged": 1.0, "promote": 1.05}

    band_toggle = {}
    for f in features:
        st = statuses.get(f["id"]) or fallback_status(f)
        wave = f["wave"]
        dur = durations.get(str(wave), 7)
        planned_start = None
        dates = {}
        if start_date_set:
            off = wave_start_offset.get(wave, 0)
            planned_start = base_date + datetime.timedelta(days=off)
            for name, frac in milestone_fracs.items():
                dates[name] = planned_start + datetime.timedelta(days=round(dur * frac))

        overall = st.get("overall") or "Not-started"
        applicable = [k for k in S_KEYS if f["subtasksApplicable"].get(k)]
        done_ct = sum(1 for k in applicable if st.get("subtasks", {}).get(k) == "Done")
        total_ct = len(applicable) or 1
        frac_done = done_ct / total_ct

        if overall == "Done":
            status_vs_plan = "On-track"
        elif overall == "Blocked":
            status_vs_plan = "Late"
        elif start_date_set and dates:
            today = datetime.date.today()
            expected_frac = 0.0
            if today >= planned_start:
                elapsed = (today - planned_start).days
                expected_frac = min(1.0, elapsed / max(dur, 1))
            if frac_done + 0.15 >= expected_frac:
                status_vs_plan = "Ahead" if frac_done > expected_frac + 0.15 else "On-track"
            elif today > dates.get("promote", today) and frac_done < 1.0:
                status_vs_plan = "Late"
            else:
                status_vs_plan = "At-risk"
        else:
            status_vs_plan = "On-track"

        row = [
            f["id"], f["name"], f"Program {f['program']}", str(wave),
            planned_start.isoformat() if planned_start else "TBD",
            dates.get("spec").isoformat() if dates.get("spec") else "TBD",
            dates.get("build").isoformat() if dates.get("build") else "TBD",
            dates.get("review").isoformat() if dates.get("review") else "TBD",
            dates.get("live").isoformat() if dates.get("live") else "TBD",
            dates.get("merged").isoformat() if dates.get("merged") else "TBD",
            dates.get("promote").isoformat() if dates.get("promote") else "TBD",
            dur, status_vs_plan, f["owner"] or "—",
        ]
        ws.append(row)
        band_toggle[ws.max_row] = wave

    waves_seen = []
    for r in range(2, ws.max_row + 1):
        w = band_toggle[r]
        if not waves_seen or waves_seen[-1] != w:
            waves_seen.append(w)
        band = WAVE_BAND_A if (len(waves_seen) % 2) else WAVE_BAND_B
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = BORDER_THIN
            if c not in (13,):
                cell.fill = band
        status_cell = ws.cell(row=r, column=13)
        status_cell.fill = STATUS_VS_PLAN_FILL.get(status_cell.value, GATE_NO_FILL)

    set_col_widths(ws, [7, 30, 10, 6, 13, 12, 14, 13, 12, 10, 10, 10, 13, 12])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    return ws


def main():
    tracker_data = load_tracker_data()
    features = tracker_data["features"]
    statuses = load_all_status()

    pillar_fills = pillar_fill_map(features)

    wb = Workbook()
    build_catalog_sheet(wb, features, pillar_fills)
    build_progress_sheet(wb, features, statuses)
    build_schedule_sheet(wb, features, statuses, tracker_data)

    wb.save(OUT_FILE)
    print(f"wrote {OUT_FILE} ({len(features)} features, sheets: {[ws.title for ws in wb.worksheets]})")


if __name__ == "__main__":
    main()
